import pandas as pd
import openpyxl
from pathlib import Path
from xlrd import XLRDError

pd.set_option('display.max_columns', 100)
pd.set_option('display.max_rows', 100)
pd.set_option('display.width', 5000)

path = Path('./pliki/').absolute()
files = [str(p) for p in path.rglob('*.xlsx')]

result = pd.DataFrame(columns=['BOM', 'ControlCabinets', 'PowerCabinets', 'PLCIntegration', 'ROBNew',
                               'ROBIntegration', 'VIBN_step_2', 'VIBN_step_3']).set_index('BOM')
hours = pd.DataFrame(columns=['BOM', 'HWDesign', 'SWDesign', 'Installation', 'StartUp', 'OLP', 'ROB']).set_index('BOM')

for file in files:
    print(file)
    if file.__contains__("W34_F54"):
        print("omijam")
        continue
    print ('dodaje')

    try:
        workbook = openpyxl.load_workbook(file, data_only=True, read_only=True)
    except (EOFError, XLRDError):
        pass
    else:

        bom = file.split("\\")[-3]
        if bom not in result.index:
            result.loc[bom] = [0, 0, 0, 0, 0, 0, 0]
            hours.loc[bom] = [0, 0, 0, 0, 0, 0]

        # VIBN flagi
        sheet = workbook["Sonstiges"]
        df = pd.DataFrame(sheet.values).iloc[:, 1:3].fillna(value=0).astype(str)
        str_vibn = "VIBN"
        str_vibn_step_2 = r"[s,S][t,T].{2,4}2"
        str_vibn_step_3 = r"[s,S][t,T].{2,4}3"

        try:
            vibn_step_2 = (df.loc[df[1].str.contains(str_vibn_step_2, na=False, case=False), [2]].iloc[0]
                           .str.contains(r"[j,J,y,Y].", na=False, case=False)).sum().sum()
        except IndexError:
            vibn_step_2 = 0

        try:
            vibn_step_3 = (df.loc[df[1].str.contains(str_vibn_step_3, na=False, case=False), [2]].iloc[0]
                           .str.contains(r"[j,J,y,Y].", na=False, case=False)).sum().sum()
        except IndexError:
            vibn_step_3 = (df.loc[df[1].str.contains(str_vibn, na=False, case=False).shift(-1).fillna(value=False), [2]]
                           .iloc[0].str.contains(r"[j,J,y,Y].", na=False, case=False)).sum().sum()

        # PLC nowe
        sheet = workbook['Schränke,UV']

        str_control_cabinet = ['BVS-Schrank', 'OC*CPC-24HP'] if sheet.cell(1,1).value == 'g' \
            else ['OVC-panel', 'OC*CPC-24HP']
        str_power_cabinet = ['Einspeiseschrank mit','Einspeisung mit'] if sheet.cell(1,1).value == 'g' \
            else ['power cabinet with', 'Power cabinet*CE']

        df = pd.DataFrame(sheet.values).iloc[:, 1:3].fillna(value=0).astype(str)

        result.at[bom, 'ControlCabinets'] += (df.loc[df[1].str.contains('|'.join(str_control_cabinet), na=False, case=False), [2]].
                                              astype(int).sum().sum())
        result.at[bom, 'PowerCabinets'] += (df.loc[df[1].str.contains('|'.join(str_power_cabinet), na=False, case=False), [2]].
                                            astype(int).sum().sum())
        #VIBN zliczanie
        if vibn_step_2 > 0:
            result.at[bom, 'VIBN_step_2'] += (df.loc[df[1].str.contains('|'.join(str_control_cabinet), na=False, case=False), [2]].
                                              astype(int).sum().sum())
        if vibn_step_3 > 0:
            result.at[bom, 'VIBN_step_3'] += (df.loc[df[1].str.contains('|'.join(str_control_cabinet), na=False, case=False), [2]].
                                              astype(int).sum().sum())

        # PLC integracja
        sheet = workbook["Integration"]
        str_plc_integration = 'Anzahl SPS`n, in die integriert wird' if sheet.cell(1,1).value == 'g' \
            else 'Number of  PLCs, which will be the integrated'
        df = pd.DataFrame(sheet.values).iloc[:, 1:3].fillna(value=0).astype(str)

        result.at[bom, 'PLCIntegration'] += (df.loc[df[1].str.contains(str_plc_integration, na=False, case=False), [2]].
                                               astype(int).sum().sum())

        #ROB nowe
        sheet = workbook["Roboter"]
        df = pd.DataFrame(sheet.values).iloc[8:, 1:3].fillna(value=0).astype(str).reset_index(drop=True)

        result.at[bom, "ROBNew"] += (df.loc[df[1].str.contains(r"\d?i?r\d+", na=False, case=False), [2]]
                                     .astype(int).sum().sum())

        #ROB integracja
        sheet = workbook["IRP Änderungen"]
        df = pd.DataFrame(sheet.values).iloc[11:, 1].fillna(value=0).astype(str).reset_index(drop=True)

        result.at[bom, 'ROBIntegration'] += df[df.str.contains(r"\d?i?r\d+", na=False, case=False)].shape[0]

        #godziny
        sheet = workbook["Gesamt"]
        df = pd.DataFrame(sheet.values).fillna(value=0).astype(str)
        str_total_hours = "Gesamt" if sheet.cell(1, 1).value == "g" else "total"

        try:
            index = df[df.iloc[:,1].str.contains(str_total_hours, na=False, case=False)].index.values[0]
        except IndexError:
            index = 38
            print("Błąd pliku!")

        df = df.iloc[index, :]

        hours.at[bom, 'HWDesign'] += float(df[55])
        hours.at[bom, 'SWDesign'] += float(df[57])
        hours.at[bom, 'Installation'] += float(df[59])
        hours.at[bom, 'StartUp'] += float(df[61])
        hours.at[bom, 'OLP'] += float(df[63])
        hours.at[bom, 'ROB'] += float(df[65])


print(result)
print(hours)
